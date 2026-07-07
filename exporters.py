"""
exporters.py — Export jobs to CSV and formatted Excel.

Excel output includes:
  - Auto-sized columns
  - Color-coded Role Type column
  - Bold headers
  - Freeze pane on row 1
  - Hyperlinked URLs
"""

import csv
import logging
from pathlib import Path
from typing import List

from models import Job

logger = logging.getLogger(__name__)

# Role type colors for Excel
ROLE_COLORS = {
    "Compliance / Risk / Fraud":    "FFD7E8",  # pink
    "EA / Business Partner":        "D7E8FF",  # blue
    "Data / Analytics":             "D7FFE8",  # green
    "Operations Analytics":         "FFF3D7",  # yellow
    "Project / Program Coordinator":"F0D7FF",  # purple
    "Other":                        "F0F0F0",  # grey
}

FIELDNAMES = [
    "Score", "Role Type", "Work Type", "Sector",
    "Title", "Company", "Location", "Source",
    "Matched Keywords", "Posted", "URL", "Description",
]


def export_csv(jobs: List[Job], path: str):
    """Write jobs to a CSV file."""
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()
        for job in jobs:
            writer.writerow(job.to_dict())
    logger.info(f"CSV saved: {path}  ({len(jobs)} jobs)")


def export_excel(jobs: List[Job], path: str):
    """Write jobs to a formatted Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("openpyxl not installed — skipping Excel export. Run: pip install openpyxl")
        return

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jobs"

    # Header row
    header_fill = PatternFill("solid", fgColor="1A365D")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_idx, field in enumerate(FIELDNAMES, 1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for row_idx, job in enumerate(jobs, 2):
        data = job.to_dict()
        role_color = ROLE_COLORS.get(job.role_type, "FFFFFF")
        row_fill = PatternFill("solid", fgColor=role_color)

        for col_idx, field in enumerate(FIELDNAMES, 1):
            value = data.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="top", wrap_text=(field == "Description"))

            # Make URL column clickable
            if field == "URL" and value:
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")

    # Auto-size columns (approximate)
    col_widths = {
        "Score": 7, "Role Type": 28, "Work Type": 18, "Sector": 22,
        "Title": 40, "Company": 20, "Location": 22, "Source": 22,
        "Matched Keywords": 40, "Posted": 22, "URL": 50, "Description": 60,
    }
    for col_idx, field in enumerate(FIELDNAMES, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(field, 20)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(FIELDNAMES))}1"
    try:
        wb.save(path)
        logger.info(f"Excel saved: {path}")

    except PermissionError:
        logger.error(f"❌ Could not save Excel to {path} because the file is currently open in Excel! Please close it and rerun.")